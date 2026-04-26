"""S0.5 Wave 2A — xlsx_adapter unit tests.

Two paths:
  * Stub fallback (openpyxl absent in the dev sandbox).
  * Real path — only exercised when openpyxl is importable (skipped otherwise).
"""

from __future__ import annotations

from io import BytesIO

import pytest

from parsers.xlsx_adapter import _safe_file_id, parse_xlsx


@pytest.fixture(autouse=True)
def _compress_gate_timeouts():  # noqa: D401
    yield


@pytest.fixture(autouse=True)
def _stub_redis_publish():
    yield None


def test_parse_xlsx_returns_parsed_file_in_stub_mode_when_dep_missing() -> None:
    """Even with the SDK absent, the adapter MUST return a valid ParsedFile."""
    # Empty bytes — never reaches openpyxl, so this test is environment-agnostic.
    parsed = parse_xlsx(b"", "compliance.xlsx")
    assert parsed.name == "compliance.xlsx"
    assert parsed.mime.endswith("spreadsheetml.sheet")
    # Either real (openpyxl raised on empty bytes — adapter caught + recorded
    # parse_error) or stub (openpyxl absent — adapter switched to stub mode).
    assert "adapter" in parsed.metadata


def test_parse_xlsx_safe_file_id_collapses_special_chars() -> None:
    assert _safe_file_id("Compliance Matrix.xlsx") == "compliance-matrix"
    assert _safe_file_id("Q&A v2.xlsx") == "q-a-v2"
    assert _safe_file_id("") == "xlsx-file"


def test_parse_xlsx_real_path_emits_table_per_sheet() -> None:
    """When openpyxl IS installed, verify the structured table output."""
    pytest.importorskip("openpyxl")

    from openpyxl import Workbook

    wb = Workbook()
    sheet = wb.active
    assert sheet is not None
    sheet.title = "Requirements"
    sheet.append(["ID", "Title", "Priority"])
    sheet.append(["R1", "Login", "MUST"])
    sheet.append(["R2", "Logout", "SHOULD"])

    second = wb.create_sheet("Compliance")
    second.append(["Framework", "Applies"])
    second.append(["HIPAA", "yes"])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    parsed = parse_xlsx(buffer.read(), "matrix.xlsx")
    assert len(parsed.tables) == 2
    assert parsed.tables[0]["caption"] == "Requirements"
    assert parsed.tables[1]["caption"] == "Compliance"
    # Markdown body contains the header pipe row.
    assert "| ID | Title | Priority |" in parsed.raw_text
    # One section per sheet.
    headings = [s["heading"] for s in parsed.sections]
    assert "Requirements" in headings
    assert "Compliance" in headings


def test_parse_xlsx_real_path_handles_empty_workbook() -> None:
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    wb = Workbook()
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    parsed = parse_xlsx(buffer.read(), "empty.xlsx")
    # Default empty Sheet; raw_text may be empty.
    assert parsed.metadata.get("adapter") == "xlsx"


def test_parse_xlsx_real_path_escapes_pipe_in_cell_text() -> None:
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    wb = Workbook()
    sheet = wb.active
    assert sheet is not None
    sheet.append(["col_a", "col_b"])
    sheet.append(["safe", "value | with pipe"])
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    parsed = parse_xlsx(buffer.read(), "pipe.xlsx")
    # Pipe is escaped so markdown table renders with the right column count.
    assert "value \\| with pipe" in parsed.raw_text
