"""XLSX → ParsedFile adapter using openpyxl.

XLSX is essential for compliance matrices: bid teams routinely receive a
"Requirements Tracker" workbook listing one requirement per row with
priority + category columns. The adapter flattens every sheet into:

  * ``raw_text`` — Markdown-pipe table per sheet, concatenated.
  * ``tables``  — one entry per sheet with the original cell grid + a sheet
                  caption so atom extractors can reference cell coords.
  * ``sections`` — one section per sheet (heading = sheet name, level=1) so
                   atom-extractor heuristics get clean section boundaries.

Cell coordinates land in ``line_range`` as ``(start_row, end_row)`` (1-based,
matches Excel's row numbering). Atom extractors use this when constructing
:class:`AtomSource` to give reviewers a click-through to the originating row.

Stub-fallback: if openpyxl is not importable (lean dev env), the function
still returns a syntactically valid ``ParsedFile`` with empty body — the
LLM stub-fallback path handles missing content gracefully.
"""

from __future__ import annotations

import logging
import re
from io import BytesIO
from typing import Any

from workflows.base import ParsedFile

logger = logging.getLogger(__name__)

_FILE_ID_SAFE_RE = re.compile(r"[^a-zA-Z0-9._-]+")


def _safe_file_id(name: str) -> str:
    """Best-effort filesystem-safe slug for ``ParsedFile.file_id``."""
    base = name.rsplit(".", 1)[0] if "." in name else name
    slug = _FILE_ID_SAFE_RE.sub("-", base).strip("-").lower()
    return slug or "xlsx-file"


def _cell_to_text(value: Any) -> str:
    """Coerce one cell value to a markdown-safe string."""
    if value is None:
        return ""
    text = str(value)
    # Pipe characters break ``| col | col |`` markdown tables — escape them.
    return text.replace("|", "\\|").replace("\n", " ").strip()


def _sheet_to_table(sheet: Any) -> dict[str, Any]:
    """Convert one openpyxl Worksheet into a structured table dict."""
    rows: list[list[str]] = []
    max_row = getattr(sheet, "max_row", 0) or 0
    max_col = getattr(sheet, "max_column", 0) or 0
    if max_row == 0 or max_col == 0:
        return {
            "caption": getattr(sheet, "title", ""),
            "rows": [],
            "max_row": 0,
            "max_col": 0,
        }
    for row in sheet.iter_rows(
        min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True
    ):
        rows.append([_cell_to_text(cell) for cell in row])
    return {
        "caption": getattr(sheet, "title", ""),
        "rows": rows,
        "max_row": max_row,
        "max_col": max_col,
    }


def _table_to_markdown(table: dict[str, Any]) -> str:
    """Render a sheet table as a markdown pipe table (best-effort)."""
    rows = table.get("rows", [])
    if not rows:
        return ""
    header = rows[0]
    body = rows[1:]
    head_line = "| " + " | ".join(header) + " |"
    sep = "| " + " | ".join(["---"] * len(header)) + " |"
    body_lines = [
        "| " + " | ".join(row + [""] * (len(header) - len(row))) + " |"
        for row in body
    ]
    return "\n".join([head_line, sep, *body_lines])


def parse_xlsx(content: bytes, source_filename: str = "spreadsheet.xlsx") -> ParsedFile:
    """Parse an XLSX byte blob into a :class:`ParsedFile`.

    Never raises on malformed input — returns a best-effort shell so the
    upstream pipeline can still emit a manifest entry + skip atom extraction
    for that file via low confidence.
    """
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError as exc:  # pragma: no cover — exercised when dep missing
        logger.warning("xlsx_adapter.stub_mode reason=openpyxl_missing err=%s", exc)
        return ParsedFile(
            file_id=_safe_file_id(source_filename),
            name=source_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            raw_text="",
            sections=[],
            tables=[],
            metadata={"adapter": "xlsx_stub"},
            size_bytes=len(content),
        )

    try:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001 — never raise on bad XLSX
        logger.warning("xlsx_adapter.parse_failed file=%s err=%s", source_filename, exc)
        return ParsedFile(
            file_id=_safe_file_id(source_filename),
            name=source_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            raw_text="",
            sections=[],
            tables=[],
            metadata={"adapter": "xlsx", "parse_error": str(exc)[:200]},
            size_bytes=len(content),
        )

    tables: list[dict[str, Any]] = []
    sections: list[dict[str, Any]] = []
    body_blocks: list[str] = []

    for sheet in workbook.worksheets:
        table = _sheet_to_table(sheet)
        tables.append(table)
        markdown = _table_to_markdown(table)
        # Section per sheet — atom extractors use heading boundaries to
        # populate AtomSource.section.
        sections.append(
            {
                "heading": table["caption"],
                "level": 1,
                "text": markdown,
                "line_range": (1, table["max_row"]),
            }
        )
        if markdown:
            body_blocks.append(f"## {table['caption']}\n\n{markdown}")

    raw_text = "\n\n".join(body_blocks)

    metadata: dict[str, str] = {
        "adapter": "xlsx",
        "sheet_count": str(len(workbook.worksheets)),
    }
    parsed = ParsedFile(
        file_id=_safe_file_id(source_filename),
        name=source_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        raw_text=raw_text,
        sections=sections,
        tables=tables,
        metadata=metadata,
        size_bytes=len(content),
    )
    logger.info(
        "xlsx_adapter.done file=%s sheets=%d rows=%d",
        source_filename,
        len(workbook.worksheets),
        sum(t["max_row"] for t in tables),
    )
    return parsed


__all__ = ["parse_xlsx"]
